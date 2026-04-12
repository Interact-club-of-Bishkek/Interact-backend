import io
import os
import random
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.conf import settings
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from django.views.generic import TemplateView
from django.db.models import Sum, Value, Q, DecimalField, Count, F, Prefetch
from django.db.models.functions import Coalesce

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from rest_framework import viewsets, generics, status, permissions
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.exceptions import PermissionDenied
from rest_framework_simplejwt.tokens import RefreshToken

# 🔥 ИСПРАВЛЕНИЕ: Оставляем только нужный импорт Document
from langchain_core.documents import Document
from langchain.prompts import PromptTemplate
from langchain_community.document_loaders import PyPDFLoader
from langchain_groq import ChatGroq
from langchain.chains.question_answering import load_qa_chain

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# --- Твои модели ---
from projects.models import Project, TeamMember
from directions.models import VolunteerDirection
from commands.models import Command
from .models import (
    AppSettings, Attendance, Volunteer, VolunteerApplication, BotAccessConfig, 
    ActivityTask, ActivitySubmission, YellowCard, ChatSession, ChatMessage, 
    MiniTeam, MiniTeamMembership, SponsorTask
)

from .serializers import (
    BulkAttendanceSerializer, VolunteerSerializer, VolunteerLoginSerializer, VolunteerRegisterSerializer,
    VolunteerApplicationSerializer, ActivityTaskSerializer, 
    ActivitySubmissionSerializer, VolunteerDirectionSerializer, CommandSerializer,
    VolunteerListSerializer, MiniTeamSerializer, SponsorTaskSerializer
)


# ---------------- АВТОРИЗАЦИЯ И ПРОФИЛЬ ----------------
class VolunteerLoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = VolunteerLoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        volunteer = serializer.validated_data.get("user") or serializer.validated_data.get("volunteer")
        
        is_responsible = VolunteerDirection.objects.filter(responsible=volunteer).exists()
        is_leader = Command.objects.filter(leader=volunteer).exists()
        
        if is_responsible or is_leader:
            if volunteer.role == 'volunteer':
                volunteer.role = 'curator'
                volunteer.is_staff = True
                volunteer.save()
            elif not volunteer.is_staff:
                volunteer.is_staff = True
                volunteer.save()

        refresh = RefreshToken.for_user(volunteer)
        
        # Исправление отношения, если у пользователя нет команд
        is_assigned = volunteer.direction.exists() or volunteer.volunteer_commands.exists()
        
        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'role': volunteer.role,
            'name': volunteer.name or volunteer.login,
            'is_assigned': is_assigned,
            'is_team_leader': is_leader,
            'is_direction_curator': is_responsible
        })

class VolunteerRegisterView(APIView):
    permission_classes = [AllowAny]
    
    def post(self, request):
        serializer = VolunteerRegisterSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            refresh = RefreshToken.for_user(user)
            return Response({
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "role": user.role,
                "is_assigned": False,
                "user": VolunteerSerializer(user, context={'request': request}).data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class VolunteerProfileView(generics.RetrieveUpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = VolunteerSerializer

    def get_object(self):
        return self.request.user

    def retrieve(self, request, *args, **kwargs):
        response = super().retrieve(request, *args, **kwargs)
        data = response.data
        
        memberships = MiniTeamMembership.objects.filter(volunteer=self.request.user)
        
        # ДОБАВЛЕНО: Явно передаем массив ролей мини-команд для фронтенда
        data['miniteam_roles'] = []
        
        if memberships.exists():
            roles = list(set([m.get_role_display() for m in memberships]))
            current_display = data.get('role_display', 'Волонтер')
            
            if data.get('role') == 'volunteer':
                data['role_display'] = f"{current_display} ({', '.join(roles)})"
                
            # ДОБАВЛЕНО: Заполняем массив нужными данными
            for m in memberships:
                data['miniteam_roles'].append({
                    'miniteam_id': m.miniteam_id,
                    'role': m.role,
                    'role_display': m.get_role_display()
                })
            
        return Response(data)

# ---------------- ЛИЧНЫЙ КАБИНЕТ ВОЛОНТЕРА ----------------
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def volunteer_direction_preferences(request):
    settings = AppSettings.get_settings()
    
    if request.method == 'GET':
        return Response({
            "is_open": settings.is_direction_selection_open,
            "preferred": list(request.user.preferred_directions.values_list('id', flat=True))
        })

    if not settings.is_direction_selection_open:
        return Response({"error": "Выбор направлений сейчас закрыт."}, status=403)

    direction_ids = request.data.get('directions', [])
    
    if len(direction_ids) > 4:
        return Response({"error": "Можно выбрать максимум 4 направления."}, status=400)

    valid_ids = VolunteerDirection.objects.filter(id__in=direction_ids).values_list('id', flat=True)
    request.user.preferred_directions.set(valid_ids)
    
    return Response({"message": "Ваши предпочтения успешно сохранены!", "saved": valid_ids})

class VolunteerActivityViewSet(viewsets.ModelViewSet):
    queryset = ActivitySubmission.objects.all()
    serializer_class = ActivitySubmissionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role in ['admin', 'bailiff_activity', 'president', 'curator']:
            return ActivitySubmission.objects.all().select_related('task').order_by('-created_at')
        return ActivitySubmission.objects.filter(volunteer=user).select_related('task').order_by('-created_at')
    
    def create(self, request, *args, **kwargs):
        settings = AppSettings.get_settings()
        if not settings.is_points_submission_open:
            return Response(
                {"error": "Прием заявок закрыт. Ожидайте дальнейших новостей"}, 
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        user = request.user
        is_staff = user.role in ['admin', 'bailiff_activity', 'president', 'curator']
        
        if not is_staff and instance.status != 'pending':
            return Response(
                {"error": "Вы не можете изменить уже обработанную заявку."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if 'points_awarded' in request.data:
            if not is_staff:
                return Response(
                    {"error": "Вам запрещено изменять баллы!"}, 
                    status=status.HTTP_403_FORBIDDEN
                )
            instance.points_awarded = request.data['points_awarded']
            instance.save()
            
        return super().partial_update(request, *args, **kwargs)

    def perform_create(self, serializer):
        command_id = self.request.data.get('command')
        direction_id = self.request.data.get('direction')
        quantity = self.request.data.get('quantity', 1)

        instance = serializer.save(
            volunteer=self.request.user,
            command_id=command_id,
            direction_id=direction_id,
            quantity=quantity
        )
        if instance.command and not instance.direction:
            instance.direction = instance.command.direction
            instance.save()
            
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        user = request.user
        
        if user.role not in ['admin', 'bailiff_activity', 'president', 'curator']:
            if instance.status != 'pending':
                return Response(
                    {"error": "Вы можете отменить только свои заявки, которые находятся в ожидании."}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class DeductPointsView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        is_leader = Command.objects.filter(leader=request.user).exists()
        if request.user.role not in ['admin', 'curator', 'president'] and not is_leader:
            return Response({"error": "Нет прав для начисления штрафов"}, status=status.HTTP_403_FORBIDDEN)

        vol_id = request.data.get('volunteer_id')
        raw_points = request.data.get('points', 0)
        reason = request.data.get('reason', 'Штраф')

        try:
            points_to_deduct = Decimal(str(raw_points))
        except (ValueError, TypeError, InvalidOperation):
            return Response({"error": "Некорректная сумма баллов"}, status=status.HTTP_400_BAD_REQUEST)

        if points_to_deduct <= 0:
            return Response({"error": "Сумма штрафа должна быть больше нуля"}, status=status.HTTP_400_BAD_REQUEST)

        volunteer = get_object_or_404(Volunteer, id=vol_id)
        
        with transaction.atomic():
            task, _ = ActivityTask.objects.get_or_create(
                title="⚠️ Штраф / Списание баллов",
                defaults={'points': 0, 'is_flexible': True}
            )
            
            ActivitySubmission.objects.create(
                volunteer=volunteer,
                task=task,
                status='approved',
                points_awarded=-points_to_deduct, 
                description=f"Списание: {reason}",
                command_id=request.data.get('command_id') 
            )
            
        return Response({"status": "success", "message": "Баллы успешно списаны"})

class DiscoveryListView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        settings = AppSettings.get_settings()
        
        all_directions = VolunteerDirection.objects.all()
        user_commands = user.volunteer_commands.all() 
        user_directions = user.direction.all()

        tasks = ActivityTask.objects.filter(
            Q(command__isnull=True) |   
            Q(command__in=user_commands) 
        ).select_related('command').distinct()

        return Response({
            "all_directions": VolunteerDirectionSerializer(all_directions, many=True).data,
            "my_direction": VolunteerDirectionSerializer(user_directions, many=True).data,
            "my_commands": CommandSerializer(user_commands, many=True).data,
            "available_tasks": ActivityTaskSerializer(tasks, many=True).data,
            "is_points_submission_open": settings.is_points_submission_open 
        })

# ---------------- ПАНЕЛЬ КУРАТОРА ----------------
class CuratorSubmissionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = ActivitySubmissionSerializer

    def get_queryset(self):
        user = self.request.user
        uid = user.id
        
        qs = ActivitySubmission.objects.select_related(
            'task', 'volunteer', 'command', 'direction'
        ).prefetch_related('volunteer__direction')

        if user.is_superuser or user.role == 'admin':
            return qs.order_by('-created_at')

        is_team_leader = Q(command__leader_id=uid)
        is_direction_curator = Q(direction__responsible_id=uid, command__isnull=True)
        is_overseer = Q(command__direction__responsible_id=uid, status='approved')

        return qs.filter(
            is_team_leader | is_direction_curator | is_overseer
        ).distinct().order_by('-created_at')

    def perform_update(self, serializer):
            serializer.save()
    
class VolunteerApplicationViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = VolunteerApplicationSerializer
    queryset = VolunteerApplication.objects.all()

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.role == 'admin':
            return VolunteerApplication.objects.all()
        return VolunteerApplication.objects.filter(
            commands__leader=user
        ).distinct()

    def perform_create(self, serializer):
        user = self.request.user
        data = serializer.validated_data

        if 'full_name' in data: user.name = data['full_name']
        if 'phone_number' in data: user.phone_number = data['phone_number']
        if 'email' in data: user.email = data['email']
        if 'direction' in data and data['direction']: user.direction.set([data['direction']])
        if 'commands' in data: user.volunteer_commands.set(data['commands'])

        is_responsible = VolunteerDirection.objects.filter(responsible=user).exists()
        is_leader = Command.objects.filter(leader=user).exists()

        if is_responsible or is_leader:
            user.role = 'curator'
            user.is_staff = True

        user.save()

class VolunteerListView(generics.ListAPIView):
    serializer_class = VolunteerListSerializer  
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        managed_commands = Command.objects.filter(Q(leader=user) | Q(direction__responsible=user))
        qs = Volunteer.objects.prefetch_related('direction', 'volunteer_commands')

        qs = qs.annotate(
            calculated_total=Coalesce(
                Sum(
                    Coalesce(
                        'submissions__points_awarded', 
                        F('submissions__task__points') * F('submissions__quantity'),
                        output_field=DecimalField()
                    ), 
                    filter=Q(submissions__status='approved')
                ),
                Value(0),
                output_field=DecimalField()
            ),
            local_points=Coalesce(
                Sum(
                    Coalesce(
                        'submissions__points_awarded', 
                        F('submissions__task__points') * F('submissions__quantity'),
                        output_field=DecimalField()
                    ), 
                    filter=Q(submissions__command__in=managed_commands, submissions__status='approved')
                ), 
                Value(0), 
                output_field=DecimalField()
            ),
            yellow_card_count=Count('yellow_cards', distinct=True)
        )
        return qs.order_by('name')

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        data_list = response.data.get('results', response.data) if isinstance(response.data, dict) else response.data
        
        if not data_list: return response

        vol_ids = [v['id'] for v in data_list]
        memberships = MiniTeamMembership.objects.filter(volunteer_id__in=vol_ids)
        
        mem_map = {}
        for m in memberships:
            if m.volunteer_id not in mem_map:
                mem_map[m.volunteer_id] = []
            mem_map[m.volunteer_id].append({
                'miniteam_id': m.miniteam_id,
                'role': m.role,
                'role_display': m.get_role_display()
            })
        
        for vol in data_list:
            vol['miniteam_roles'] = mem_map.get(vol['id'], [])
            
        return response

class RemoveVolunteerFromCommandView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        command = get_object_or_404(Command, pk=pk)
        
        if command.leader != request.user and not request.user.is_superuser:
            return Response({"error": "Нет прав"}, status=status.HTTP_403_FORBIDDEN)

        vol_id = request.data.get('volunteer_id')
        if not vol_id:
            return Response({"error": "ID волонтера не передан"}, status=status.HTTP_400_BAD_REQUEST)

        volunteer = get_object_or_404(Volunteer, pk=vol_id)
        command.volunteers.remove(volunteer)
        
        return Response({"status": "success", "message": f"{volunteer.name} удален из команды"})


class VolunteerViewSet(viewsets.ModelViewSet):
    serializer_class = VolunteerSerializer
    permission_classes = [IsAuthenticated]
    queryset = Volunteer.objects.all()

    def get_queryset(self):
        user = self.request.user
        is_leader = Command.objects.filter(leader=user).exists()
        
        if user.is_staff or user.role in ['admin', 'curator'] or is_leader:
            return Volunteer.objects.all().order_by('-date_joined')
        
        return Volunteer.objects.filter(id=user.id)
    
class AttendanceViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def month_journal(self, request):
        direction_id = request.query_params.get('direction_id')
        month_str = request.query_params.get('month') 
        
        if not direction_id or not month_str:
            return Response({"error": "Нужны direction_id и month"}, status=400)

        try:
            year, month = map(int, month_str.split('-'))
        except ValueError:
            return Response({"error": "Неверный формат даты"}, status=400)

        volunteers = Volunteer.objects.filter(
            direction__id=direction_id, role='volunteer'
        ).order_by('name')
        
        logs = Attendance.objects.filter(
            direction_id=direction_id, date__year=year, date__month=month
        ).order_by('date')

        existing_dates = sorted(list(set([log.date.strftime('%Y-%m-%d') for log in logs])))

        journal_map = {}
        for log in logs:
            vid = log.volunteer_id
            d_str = log.date.strftime('%Y-%m-%d')
            if vid not in journal_map: journal_map[vid] = {}
            journal_map[vid][d_str] = log.status

        vol_list = []
        for vol in volunteers:
            name = vol.name or vol.login
            parts = name.split()
            initials = (parts[0][0] + (parts[1][0] if len(parts)>1 else "")).upper()[:2]
            
            vol_list.append({
                "id": vol.id,
                "name": name,
                "initials": initials,
                "records": journal_map.get(vol.id, {})
            })
            
        return Response({
            "dates": existing_dates,
            "volunteers": vol_list
        })

    @action(detail=False, methods=['post'])
    def mark_bulk(self, request):
        data = request.data
        direction_id = data.get('direction_id')
        records = data.get('records', [])

        if request.user.role not in ['bailiff_activity', 'admin', 'curator', 'president']:
            return Response({"error": "Нет прав"}, status=403)

        saved_count = 0
        with transaction.atomic():
            for item in records:
                date_str = item.get('date')
                vol_id = item.get('volunteer_id')
                status = item.get('status')

                if not date_str or not vol_id: continue

                if not status: 
                    Attendance.objects.filter(
                        volunteer_id=vol_id, direction_id=direction_id, date=date_str
                    ).delete()
                else:
                    Attendance.objects.update_or_create(
                        volunteer_id=vol_id, direction_id=direction_id, date=date_str,
                        defaults={'status': status, 'marked_by': request.user}
                    )
                saved_count += 1
            
        return Response({"message": "Сохранено"})

    @action(detail=False, methods=['get'])
    def stats_by_month(self, request):
        directions = VolunteerDirection.objects.all()
        data = []

        for current_dir in directions:
            vols = Volunteer.objects.filter(
                direction=current_dir, 
                role='volunteer'
            ).annotate(
                total_score=Coalesce(
                    Sum(
                        Coalesce(
                            'submissions__points_awarded', 
                            F('submissions__task__points') * F('submissions__quantity'),
                            output_field=DecimalField()
                        ), 
                        filter=Q(submissions__status='approved')
                    ),
                    Value(0),
                    output_field=DecimalField()
                )
            ).prefetch_related(
                'submissions', 
                'submissions__task',
                'submissions__command',
                'submissions__direction'
            )

            vol_list = []
            for v in vols:
                score = float(v.total_score)
                approved_subs = [s for s in v.submissions.all() if s.status == 'approved']
                tasks_data = []
                for sub in approved_subs:
                    qty = getattr(sub, 'quantity', 1) 
                    
                    if sub.points_awarded is not None:
                        p = float(sub.points_awarded)
                    else:
                        p = float(sub.task.points) * qty if sub.task else 0.0

                    tasks_data.append({
                        "id": sub.id, 
                        "title": sub.task.title if sub.task else "Без названия",
                        "points": p,
                        "date": sub.created_at.strftime('%Y-%m-%d'),
                        "description": sub.description,
                        "command_title": sub.command.title if sub.command else None,
                        "direction_name": sub.direction.name if sub.direction else None,
                        "quantity": qty
                    })

                vol_list.append({
                    "id": v.id,
                    "name": v.name or v.login,
                    "score": score,
                    "tasks": tasks_data
                })

            if vol_list:
                data.append({
                    "direction_id": current_dir.id,
                    "direction_name": current_dir.name,
                    "volunteers": vol_list
                })

        return Response(data)
    
    @action(detail=False, methods=['get'])
    def download_all_stats_excel(self, request):
        wb = Workbook()
        wb.remove(wb.active) 

        directions = VolunteerDirection.objects.all()

        DIR_COLORS = {
            'СС':   {'tab': '0070C0', 'main': 'C9DAF8', 'light': 'E8F0FE'},
            'ЭКО':  {'tab': '00B050', 'main': 'D9EAD3', 'light': 'F0F4EC'},
            'ОНКО': {'tab': '7030A0', 'main': 'E4DFEC', 'light': 'F3F0F5'},
            'ЛОВЗ': {'tab': 'FFC000', 'main': 'FFF2CC', 'light': 'FFF9E6'},
            'КЦ':   {'tab': 'FF0000', 'main': 'FADAD8', 'light': 'FDEDED'},
            'МС':   {'tab': 'E26B0A', 'main': 'FCE4D6', 'light': 'FEF0E6'},
            'ДП':   {'tab': '31869B', 'main': 'D0E0E3', 'light': 'E9F1F2'},
            'ДД':   {'tab': '1F497D', 'main': 'CFE2F3', 'light': 'EBF1F7'},
        }

        thin_side = Side(style='thin')
        medium_side = Side(style='medium')
        
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        box_left_border = Border(left=medium_side, top=medium_side, bottom=medium_side, right=thin_side)
        box_right_border = Border(right=medium_side, top=medium_side, bottom=medium_side, left=thin_side)

        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")

        def get_category(sub):
            if hasattr(sub, 'task') and sub.task:
                if hasattr(sub.task, 'command') and sub.task.command:
                    return getattr(sub.task.command, 'name', getattr(sub.task.command, 'title', 'Команда'))
            return "Общие задания"

        for current_dir in directions:
            vols = Volunteer.objects.filter(
                direction=current_dir, role='volunteer'
            ).annotate(
                total_score=Coalesce(
                    Sum(
                        Coalesce('submissions__points_awarded', F('submissions__task__points') * F('submissions__quantity'), output_field=DecimalField()), 
                        filter=Q(submissions__status='approved')
                    ), Value(0), output_field=DecimalField()
                )
            ).prefetch_related(
                Prefetch('submissions', queryset=ActivitySubmission.objects.filter(status='approved').select_related('task').order_by('created_at'))
            ).order_by('name')

            if not vols.exists():
                continue

            sheet_title = str(current_dir.name)[:31]
            ws = wb.create_sheet(title=sheet_title)

            dir_name_upper = current_dir.name.upper()
            tab_color = "A6A6A6"  
            main_color = "D9D9D9" 
            light_color = "F2F2F2"
            
            for key, colors_data in DIR_COLORS.items():
                if key in dir_name_upper:
                    tab_color = colors_data['tab']
                    main_color = colors_data['main']
                    light_color = colors_data['light']
                    break
            
            main_fill = PatternFill(start_color=main_color, end_color=main_color, fill_type="solid")
            light_fill = PatternFill(start_color=light_color, end_color=light_color, fill_type="solid")
            ws.sheet_properties.tabColor = tab_color

            start_row = 2

            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 3 
            ws.row_dimensions[start_row].height = 25

            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
            c_dir = ws.cell(row=start_row, column=1, value=current_dir.name.upper())
            c_dir.fill, c_dir.font, c_dir.alignment = main_fill, bold_font, center_aligned
            ws.cell(row=start_row, column=1).border = box_left_border
            ws.cell(row=start_row, column=2).border = box_right_border

            for idx_v, vol in enumerate(vols):
                row_idx = start_row + 1 + idx_v
                
                c_vname = ws.cell(row=row_idx, column=1, value=vol.name or vol.login)
                c_vname.fill, c_vname.alignment, c_vname.border = light_fill, left_aligned, thin_border
                
                c_vscore = ws.cell(row=row_idx, column=2, value=round(float(vol.total_score), 2))
                c_vscore.fill, c_vscore.alignment, c_vscore.border, c_vscore.font = main_fill, center_aligned, thin_border, bold_font

            total_dir_score = sum(float(v.total_score) for v in vols)
            avg_dir_score = round(total_dir_score / len(vols), 2) if len(vols) > 0 else 0
            
            avg_row_idx = start_row + 1 + len(vols)
            c_avg_lbl = ws.cell(row=avg_row_idx, column=1, value="Среднее арифметическое баллов")
            c_avg_lbl.fill, c_avg_lbl.font, c_avg_lbl.alignment, c_avg_lbl.border = light_fill, bold_font, left_aligned, box_left_border
            
            c_avg_val = ws.cell(row=avg_row_idx, column=2, value=avg_dir_score)
            c_avg_val.fill, c_avg_val.font, c_avg_val.alignment, c_avg_val.border = main_fill, bold_font, center_aligned, box_right_border

            task_start_col = 8 

            for i, vol in enumerate(vols):
                v_col = task_start_col + (i * 3) 
                
                ws.column_dimensions[get_column_letter(v_col)].width = 42
                ws.column_dimensions[get_column_letter(v_col+1)].width = 9
                ws.column_dimensions[get_column_letter(v_col+2)].width = 3 

                ws.merge_cells(start_row=start_row, start_column=v_col, end_row=start_row, end_column=v_col+1)
                
                cell_left = ws.cell(row=start_row, column=v_col, value=vol.name or vol.login)
                cell_left.fill, cell_left.font, cell_left.alignment = main_fill, bold_font, center_aligned
                cell_left.border = box_left_border 
                ws.cell(row=start_row, column=v_col+1).border = box_right_border 
                ws.cell(row=start_row, column=v_col+1).fill = main_fill 

                task_row_idx = start_row + 1
                approved_subs = vol.submissions.all()
                
                subs_by_cat = {}
                for sub in approved_subs:
                    cat = get_category(sub)
                    if cat not in subs_by_cat:
                        subs_by_cat[cat] = []
                    subs_by_cat[cat].append(sub)
                
                sorted_cats = sorted(subs_by_cat.keys(), key=lambda x: (x == "Общие задания", x))
                
                for cat in sorted_cats:
                    ws.merge_cells(start_row=task_row_idx, start_column=v_col, end_row=task_row_idx, end_column=v_col+1)
                    c_cat = ws.cell(row=task_row_idx, column=v_col, value=str(cat).upper())
                    c_cat.fill, c_cat.font, c_cat.alignment, c_cat.border = main_fill, bold_font, center_aligned, thin_border
                    ws.cell(row=task_row_idx, column=v_col+1).border = thin_border
                    ws.cell(row=task_row_idx, column=v_col+1).fill = main_fill
                    task_row_idx += 1
                    
                    for sub in subs_by_cat[cat]:
                        task_title = sub.task.title if sub.task else "Без названия"
                        qty = getattr(sub, 'quantity', 1)
                        points = float(sub.points_awarded) if sub.points_awarded is not None else (float(sub.task.points) * qty if sub.task else 0.0)
                        
                        c_task = ws.cell(row=task_row_idx, column=v_col, value=task_title)
                        c_task.border, c_task.alignment, c_task.fill = thin_border, left_aligned, light_fill
                        
                        c_pts = ws.cell(row=task_row_idx, column=v_col+1, value=round(points, 2))
                        c_pts.alignment, c_pts.border, c_pts.fill = center_aligned, thin_border, main_fill
                        
                        task_row_idx += 1
                
                c_blank = ws.cell(row=task_row_idx, column=v_col, value="")
                c_blank.border, c_blank.fill = box_left_border, main_fill
                
                c_sum = ws.cell(row=task_row_idx, column=v_col+1, value=round(float(vol.total_score), 2))
                c_sum.border, c_sum.font, c_sum.alignment, c_sum.fill = box_right_border, bold_font, center_aligned, main_fill

        if not wb.sheetnames:
            wb.create_sheet(title="Пусто")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename="Activity_Stats.xlsx", content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
class BailiffPanelView(TemplateView):
    template_name = "volunteers/bailiff_panel.html"

class EquityViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def board(self, request):
        direction_id = request.query_params.get('direction_id')
        if not direction_id:
            return Response({"error": "Нужен direction_id"}, status=400)

        volunteers = Volunteer.objects.filter(
            direction__id=direction_id, 
            role='volunteer'
        ).order_by('name')

        data = []
        for vol in volunteers:
            cards = YellowCard.objects.filter(volunteer=vol).order_by('date_issued')
            card_data = [{"id": c.id, "date": c.date_issued, "reason": c.reason} for c in cards]
            
            data.append({
                "id": vol.id,
                "name": vol.name or vol.login,
                "cards": card_data,
                "count": len(card_data)
            })

        return Response(data)

    @action(detail=False, methods=['post'])
    def toggle_card(self, request):
        if request.user.role not in ['equity_officer', 'admin', 'curator', 'president']:
             return Response({"error": "Нет прав"}, status=403)

        vol_id = request.data.get('volunteer_id')
        action_type = request.data.get('action') 

        try:
            volunteer = Volunteer.objects.get(id=vol_id)
        except Volunteer.DoesNotExist:
            return Response({"error": "Волонтер не найден"}, status=404)

        current_count = YellowCard.objects.filter(volunteer=volunteer).count()

        if action_type == 'add':
            if current_count >= 4:
                return Response({"error": "Максимум 4 карточки!"}, status=400)
            
            YellowCard.objects.create(
                volunteer=volunteer,
                issued_by=request.user,
                reason=request.data.get('reason', 'Нарушение')
            )
            return Response({"message": "Карточка выдана", "new_count": current_count + 1})

        elif action_type == 'remove':
            if current_count == 0:
                return Response({"error": "У волонтера нет карточек"}, status=400)
            
            last_card = YellowCard.objects.filter(volunteer=volunteer).last()
            if last_card:
                last_card.delete()
            
            return Response({"message": "Карточка снята", "new_count": current_count - 1})

        return Response({"error": "Неверное действие"}, status=400)
    
class EquityPanelView(TemplateView):
    template_name = "volunteers/equity_panel.html"

# ---------------- PDF ГЕНЕРАЦИЯ ----------------
class DownloadPDFBase(APIView):
    permission_classes = [AllowAny]

    def get_pdf_response(self, volunteers, title, filename):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        font_path = os.path.join(settings.BASE_DIR, 'FreeSans.ttf')
        font_name = 'FreeSans' if os.path.exists(font_path) else 'Helvetica'
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont('FreeSans', font_path))

        elements.append(Paragraph(title, ParagraphStyle('T', fontName=font_name, fontSize=18, alignment=1)))
        
        data = [['№', 'ФИО', 'Телефон']]
        for i, v in enumerate(volunteers):
            name = v.full_name if v.full_name else "Не указано"
            phone = v.phone_number if v.phone_number else "-"
            data.append([str(i+1), name, phone])
        
        t = Table(data, colWidths=[30, 300, 150])
        t.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), font_name),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename=filename)

class DownloadInterviewScheduleView(DownloadPDFBase):
    def get(self, request):
        vols = VolunteerApplication.objects.filter(status='interview').order_by('full_name')
        return self.get_pdf_response(vols, "Расписание собеседований", "Interviews.pdf")

class DownloadAcceptedNamesView(DownloadPDFBase):
    def get(self, request):
        vols = VolunteerApplication.objects.filter(status='accepted').order_by('full_name')
        return self.get_pdf_response(vols, "Принятые волонтеры", "Accepted.pdf")

# ---------------- HTML VIEWS ----------------
class VolunteerColumnsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from .models import VolunteerApplication
        columns = {
            "submitted": VolunteerApplicationSerializer(VolunteerApplication.objects.filter(status='submitted'), many=True).data,
            "interview": VolunteerApplicationSerializer(VolunteerApplication.objects.filter(status='interview'), many=True).data,
            "accepted": VolunteerApplicationSerializer(VolunteerApplication.objects.filter(status='accepted'), many=True).data,
        }
        return Response(columns)

# ==========================================
# ЛОГИКА ПРИСТАВА БАЗ (РАСПРЕДЕЛЕНИЕ)
# ==========================================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_auto_distribution(request):
    if request.user.role not in ['bailiff_base', 'admin', 'president']:
        return Response({"error": "Нет прав"}, status=403)

    volunteers = list(Volunteer.objects.filter(is_active=True, role='volunteer'))
    directions = list(VolunteerDirection.objects.all())
    
    if not directions:
        return Response({"error": "Нет направлений в базе"}, status=400)

    target_capacity = (len(volunteers) // len(directions)) + 1 
    
    direction_counts = {d.id: 0 for d in directions}
    final_mapping = {} 
    unassigned_volunteers = volunteers[:]
    
    random.shuffle(unassigned_volunteers)

    for priority_index in range(4):
        still_unassigned = []
        for vol in unassigned_volunteers:
            prefs = list(vol.preferred_directions.all()) 
            
            assigned = False
            if priority_index < len(prefs):
                chosen_dir = prefs[priority_index]
                if direction_counts[chosen_dir.id] < target_capacity:
                    final_mapping[vol.id] = chosen_dir.id
                    direction_counts[chosen_dir.id] += 1
                    assigned = True
            
            if not assigned:
                still_unassigned.append(vol)
        
        unassigned_volunteers = still_unassigned

    for vol in unassigned_volunteers:
        emptiest_dir_id = min(direction_counts, key=direction_counts.get)
        final_mapping[vol.id] = emptiest_dir_id
        direction_counts[emptiest_dir_id] += 1

    distribution_result = []
    all_vols = Volunteer.objects.filter(is_active=True, role='volunteer')
    for v in all_vols:
        distribution_result.append({
            "volunteer_id": v.id,
            "volunteer_name": v.name or v.login,
            "assigned_direction_id": final_mapping.get(v.id)
        })

    return Response({
        "distribution": distribution_result, 
        "counts": direction_counts
    })

@api_view(['GET'])
@permission_classes([AllowAny]) 
def get_app_settings(request):
    settings = AppSettings.get_settings()
    return Response({
        "is_registration_open": settings.is_registration_open,
        "is_direction_selection_open": settings.is_direction_selection_open,
        "is_points_submission_open": settings.is_points_submission_open 
    })

@api_view(['POST'])
@permission_classes([AllowAny])
def ai_pdf_chat(request):
    user_text = request.data.get('message')
    session_id = request.data.get('session_id')

    if not user_text or not session_id:
        return Response({"error": "Пустое сообщение или нет сессии"}, status=400)

    session, _ = ChatSession.objects.get_or_create(session_id=session_id)
    ChatMessage.objects.create(session=session, sender='user', text=user_text)

    try:
        pdf_path = os.path.join(settings.BASE_DIR, 'rules.pdf')
        loader = PyPDFLoader(pdf_path)
        # 🔥 ИСПРАВЛЕНИЕ: изменил переменную, чтобы не было конфликта с импортом LangChain
        loaded_docs = loader.load()

        active_projects = Project.objects.filter(is_archived=False).order_by('-time_start')[:5] 
        projects_info = "=== АКТУАЛЬНЫЕ ПРОЕКТЫ ИЗ БАЗЫ ДАННЫХ: ===\n"
        if active_projects.exists():
            for p in active_projects:
                date_str = p.time_start.strftime("%d.%m.%Y") if p.time_start else "Скоро"
                projects_info += f"- [{p.name}](/projects/{p.slug}/) (Дата: {date_str}, Категория: {p.get_category_display()})\n"
        else:
            projects_info += "В данный момент активных проектов нет.\n"

        loaded_docs.append(Document(page_content=projects_info, metadata={"source": "db_projects"}))

        team_members = TeamMember.objects.filter(is_active=True).order_by('order')
    
        team_info = "\n=== КОМАНДА И РУКОВОДСТВО КЛУБА: ===\n"
        if team_members.exists():
            for member in team_members:
                info = f"- {member.full_name}: {member.position}"
                if member.description:
                    info += f" ({member.description})"
                team_info += info + "\n"
        else:
            team_info += "Информация о членах команды временно недоступна."

        loaded_docs.append(Document(page_content=team_info, metadata={"source": "db_team"}))

        groq_api_key = os.environ.get("GROQ_API_KEY")
        llm = ChatGroq(
            api_key=groq_api_key,
            model_name="llama-3.1-8b-instant", 
            temperature=0.3,
            base_url="https://icy-dust-9f56.mamadalievmaruf740.workers.dev"
        )

        prompt_template = """Ты — дружелюбный и современный ИИ-помощник Interact Club of Bishkek.
        Опираясь на предоставленный контекст, ответь на вопрос пользователя.

        ТВОИ СТРОГИЕ ПРАВИЛА:
        1. КРАСОТА: Отвечай очень красиво и структурированно. Используй абзацы, списки и много эмодзи. 🌟
        2. О КЛУБЕ И ПРАВИЛАХ: Используй данные из PDF-файла.
        3. ПРОЕКТЫ: Если спрашивают про дела или мероприятия, бери данные из "АКТУАЛЬНЫЕ ПРОЕКТЫ".
        4. КОМАНДА: Если спрашивают "кто президент", "кто в команде", "кто лидеры", используй раздел "КОМАНДА И РУКОВОДСТВО КЛУБА".
        5. ССЫЛКИ: Выводи ссылки на проекты ровно в том виде (Markdown), в котором они даны в контексте.
        6. ЕСЛИ НЕ ЗНАЕШЬ — ЧЕСТНО СКАЖИ, ЧТО НЕ ЗНАЕШЬ. НЕ ВЫДУМЫВАЙ ИНФОРМАЦИЮ.
        7. ЯЗЫК: Отвечай на том же языке, на котором задан вопрос (русский или английский).
        8. ЭМОЦИИ: Будь максимально дружелюбным, позитивным и вдохновляющим. Добавляй эмодзи, чтобы сделать ответ живым и теплым. 😊✨
        9. Информация о нас и нашей команде по ссылке https://interact-club.kg/about/
        10. Проекты находятся по ссылке https://interact-club.kg/projects/
        11. Для спонсортсва страница https://interact-club.kg/sponsorship/
        12. Для того чтобы стать нашим волонтером, страница https://interact-club.kg/volunteer/

        Контекст:
        {context}

        Вопрос: {question}
        Красивый ответ:"""

        PROMPT = PromptTemplate(template=prompt_template, input_variables=["context", "question"])
        
        chain = load_qa_chain(llm, chain_type="stuff", prompt=PROMPT)
        result = chain.invoke({"input_documents": loaded_docs, "question": user_text})
        answer = result["output_text"]

        ChatMessage.objects.create(session=session, sender='ai', text=answer)
        return Response({"answer": answer})

    except Exception as e:
        print(f"🔴 ОШИБКА: {e}")
        return Response({"error": str(e)}, status=500)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def apply_distribution(request):
    if request.user.role not in ['bailiff_base', 'admin', 'president']:
        return Response({"error": "Нет прав"}, status=403)

    if request.method == 'GET':
        volunteers = Volunteer.objects.filter(is_active=True, role='volunteer')
        dist = []
        for vol in volunteers:
            if vol.draft_direction:
                dir_id = vol.draft_direction.id
            else:
                dir_id = vol.direction.first().id if vol.direction.exists() else ''
                
            dist.append({
                "volunteer_id": vol.id,
                "volunteer_name": vol.name or vol.login,
                "assigned_direction_id": dir_id
            })
        return Response({"distribution": dist})

    action = request.data.get('action') 
    mapping = request.data.get('mapping', [])

    if action not in ['save_only', 'distribute_and_reset']:
        return Response({"error": "Неизвестное действие"}, status=400)

    with transaction.atomic():
        for item in mapping:
            vol_id = item.get('vol_id')
            dir_id = item.get('dir_id')
            
            try:
                vol = Volunteer.objects.get(id=vol_id)
            except Volunteer.DoesNotExist:
                continue

            if action == 'save_only':
                vol.draft_direction_id = dir_id if dir_id else None
                vol.save()
                
            elif action == 'distribute_and_reset':
                vol.direction.clear()
                if dir_id:
                    vol.direction.add(dir_id)
                
                vol.draft_direction = None

                vol.point = 0
                vol.preferred_directions.clear() 
                vol.submissions.all().delete() 
                vol.save()
                
    if action == 'distribute_and_reset':
        settings = AppSettings.get_settings()
        settings.is_direction_selection_open = False
        settings.save()

    return Response({"message": f"Успешно выполнено: {action}"})

class MiniTeamViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = MiniTeamSerializer

    def get_queryset(self):
        user = self.request.user
        
        if user.role in ['admin', 'president']:
            return MiniTeam.objects.all()
            
        return MiniTeam.objects.filter(
            Q(direction__responsible=user) | 
            Q(command__leader=user) |
            Q(memberships__volunteer=user)
        ).distinct()

    def perform_create(self, serializer):
        user = self.request.user
        user_command = Command.objects.filter(leader=user).first()
        user_direction = VolunteerDirection.objects.filter(responsible=user).first()

        miniteam = serializer.save(command=user_command, direction=user_direction)
        leader_id = self.request.data.get('leader')
        
        if leader_id:
            volunteer = get_object_or_404(Volunteer, id=leader_id)
            MiniTeamMembership.objects.create(
                miniteam=miniteam,
                volunteer=volunteer,
                role='mini_curator',
                assigned_by=user
            )

    @action(detail=True, methods=['post'])
    def assign_member(self, request, pk=None):
        miniteam = self.get_object()
        user = request.user
        
        is_parent_curator = (miniteam.direction and miniteam.direction.responsible == user) or \
                            (miniteam.command and miniteam.command.leader == user)
        is_mini_curator = MiniTeamMembership.objects.filter(
            miniteam=miniteam, volunteer=user, role='mini_curator'
        ).exists()
        
        if not (is_parent_curator or is_mini_curator or user.role in ['admin', 'president']):
            raise PermissionDenied("У вас нет прав управлять составом этой мини-команды")

        vol_id = request.data.get('volunteer_id')
        role = request.data.get('role', 'member')

        if role not in dict(MiniTeamMembership.ROLE_CHOICES).keys():
            return Response({"error": "Недопустимая роль"}, status=400)

        volunteer = get_object_or_404(Volunteer, pk=vol_id)
        
        membership, created = MiniTeamMembership.objects.update_or_create(
            miniteam=miniteam, 
            volunteer=volunteer,
            defaults={'role': role, 'assigned_by': user}
        )
        
        action_text = "добавлен" if created else "обновлен"
        return Response({"message": f"Волонтер {volunteer.name} {action_text} с ролью {membership.get_role_display()}"})
    
    @action(detail=True, methods=['post'])
    def remove_member(self, request, pk=None):
        miniteam = self.get_object()
        user = request.user
        
        is_parent_curator = (miniteam.direction and miniteam.direction.responsible == user) or \
                            (miniteam.command and miniteam.command.leader == user)
        is_mini_curator = MiniTeamMembership.objects.filter(
            miniteam=miniteam, volunteer=user, role='mini_curator'
        ).exists()
        
        if not (is_parent_curator or is_mini_curator or user.role in ['admin', 'president']):
            raise PermissionDenied("У вас нет прав управлять составом этой мини-команды")

        vol_id = request.data.get('volunteer_id')
        role_to_remove = request.data.get('role')

        if not vol_id or not role_to_remove:
            return Response({"error": "ID волонтера и роль обязательны для передачи"}, status=status.HTTP_400_BAD_REQUEST)

        deleted, _ = MiniTeamMembership.objects.filter(
            miniteam=miniteam, 
            volunteer_id=vol_id,
            role=role_to_remove 
        ).delete()

        if deleted:
            return Response({"message": "Роль успешно удалена"})
        else:
            return Response({"error": "У волонтера нет такой роли в этой мини-команде"}, status=status.HTTP_404_NOT_FOUND)

class SponsorTaskViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = SponsorTaskSerializer

    def get_queryset(self):
        user = self.request.user
        
        if user.role in ['admin', 'president']:
            return SponsorTask.objects.all().order_by('-created_at')
            
        return SponsorTask.objects.filter(
            Q(miniteam__memberships__volunteer=user, miniteam__memberships__role__in=['basist', 'mini_curator']) |
            Q(miniteam__direction__responsible=user) |
            Q(miniteam__command__leader=user) |
            Q(assigned_volunteer=user)
        ).distinct().order_by('-created_at')

    def perform_create(self, serializer):
        user = self.request.user
        miniteam_id = self.request.data.get('miniteam')
        miniteam = get_object_or_404(MiniTeam, id=miniteam_id)
        
        is_basist = MiniTeamMembership.objects.filter(miniteam=miniteam, volunteer=user, role='basist').exists()
        is_parent_curator = (miniteam.direction and miniteam.direction.responsible == user) or \
                            (miniteam.command and miniteam.command.leader == user)
                            
        if not (is_basist or is_parent_curator or user.role in ['admin', 'president']):
            raise PermissionDenied("Только Базист или куратор может добавлять спонсоров в базу.")
            
        serializer.save()

    @action(detail=True, methods=['patch'])
    def update_status(self, request, pk=None):
        task = self.get_object()
        user = request.user
        
        is_assigned = (task.assigned_volunteer == user)
        is_basist = MiniTeamMembership.objects.filter(
            miniteam=task.miniteam, volunteer=user, role__in=['basist', 'mini_curator']
        ).exists()
        
        if not (is_assigned or is_basist or user.role in ['admin', 'president']):
            raise PermissionDenied("У вас нет прав изменять статус этого спонсора")

        new_status = request.data.get('status')
        comment = request.data.get('comment')

        if new_status:
            if new_status not in dict(SponsorTask.STATUS_CHOICES).keys():
                return Response({"error": "Неверный статус"}, status=400)
            task.status = new_status
            
        if comment is not None:
            task.comment = comment
            
        task.save()
        
        return Response({
            "message": "Данные по спонсору успешно обновлены", 
            "status": task.get_status_display(),
            "comment": task.comment
        })

class CuratorPanelView(TemplateView): template_name = "volunteers/curator_panel.html"
class VolunteerCabinetView(TemplateView): template_name = "volunteers/page_volunteers/volunteer_cabinet.html"
class BailiffBasePanelView(TemplateView): template_name = "volunteers/bailiff_base_panel.html"
class LoginPageView(TemplateView): template_name = "volunteers/login.html"
class VolunteerBoardView(TemplateView): template_name = "volunteers/columns.html"